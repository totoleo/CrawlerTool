# -*- coding: utf-8 -*-
__author__ = 'hipponensis'

"""年度: 电影名, 评分, 评价人数, 制片地区/类型/上映年份/导演/主演, 链接, 前5/10/20张图片下载"""

import os
import sys
import time
import urllib.request
import urllib.error
import urllib.parse
import requests
import shutil
import gzip
import numpy as np
import re
from bs4 import BeautifulSoup
from multiprocessing.dummy import Pool as ThreadPool
from openpyxl import Workbook


class DoubanMoviesCrawler(object):
    def __init__(self, topics):
        self.topics = topics
        self.headers = [
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:40.0) Gecko/20100101 Firefox/40.0', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36', 'Accept-Encoding': 'gzip'}
        ]

    def crawler(self, topic_id, topic_name):
        url = 'http://www.douban.com/j/tag/items?start=0&limit=1000000&topic_id=' + topic_id + '&topic_name=' + topic_name + '&mod=movie'
        movie_list = []
        try:
            randhds = self.headers[np.random.randint(0, len(self.headers)-1)]
            req = urllib.request.Request(url, headers=randhds)
            resp = urllib.request.urlopen(req)
            try:
                g = gzip.GzipFile(mode="rb", fileobj=resp)
                source_code = g.read()
            except:
                source_code = resp.read()
            source_code = eval(source_code.decode('utf-8', 'ignore'))
            plain_text = str(source_code['html'])
            total = int(source_code['total'])
            plain_text = plain_text.replace('\\', '')
            plain_text = plain_text.replace('\n', '')
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(e)
            plain_text = ''

        list_info = re.split(r'<dl>', plain_text)[1:]

        if list_info is None or len(list_info) <= 1:
            print(list_info)
            sys.exit(0)

        for movie_info in list_info:
            soup = BeautifulSoup(movie_info, 'html.parser')
            movie_url = soup.find('a').get('href')
            title = soup.find('a', class_="title").get_text()
            desc = soup.find('div', class_="desc").get_text()
            desc = desc.strip(' ')
            desc = desc.strip('\n')
            try:
                rating = soup.find('span', class_="rating_nums").get_text()
            except:
                rating = '0.0'
            try:
                if rating != '0.0':
                    votes = self.crawler_votes(movie_url)
                    votes = votes.replace('\\xe4\\xba\\xba\\xe8\\xaf\\x84\\xe4\\xbb\\xb7', '')
                else:
                    votes = '0'
            except:
                votes = '0'
            if int(votes) >= 10000:
                limit = 20
            elif 1000 <= int(votes) < 10000:
                limit = 10
            elif 500 <= int(votes) < 1000:
                limit = 5
            else:
                limit = 0
            movie_list.append([title, rating, votes, desc, movie_url])
            if not self.exist_file('images/movies', title + '1.jpg'):
                self.crawler_images(movie_url, title, limit)
        return movie_list
    
    def crawler_votes(self, movie_url):
        time.sleep(np.random.rand()*5)
        try:
            req = urllib.request.Request(movie_url, headers=self.headers[np.random.randint(0, len(self.headers) - 1)])
            resp = urllib.request.urlopen(req)
            try:
                g = gzip.GzipFile(mode="rb", fileobj=resp)
                source_code = g.read()
            except:
                source_code = resp.read()
            plain_text = str(source_code)
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(e, movie_url)
            plain_text = ''
        soup = BeautifulSoup(plain_text, 'html.parser')
        votes = soup.find(id="interest_sectl").find('a', class_="rating_people").get_text()
        return votes

    def crawler_images(self, movie_url, title, limit):
        if limit != 0:
            # http://movie.douban.com/subject/25897313/all_photos
            all_img_url = movie_url.replace('?from=tag', 'all_photos')
            # http://img1.doubanio.com/view/photo/albumicon/public/p2195889027.jpg
            try:
                req = urllib.request.Request(all_img_url, headers=self.headers[np.random.randint(0, len(self.headers) - 1)])
                resp = urllib.request.urlopen(req)
                try:
                    g = gzip.GzipFile(mode="rb", fileobj=resp)
                    source_code = g.read()
                except:
                    source_code = resp.read()
                plain_text = str(source_code)
            except (urllib.error.HTTPError, urllib.error.URLError) as e:
                print(e, all_img_url)
                plain_text = ''
            soup = BeautifulSoup(plain_text, 'html.parser')
            soup_results = soup.find('div', class_='mod').find('ul', class_='pic-col5').find_all('img', limit=limit)
            img_urls = []
            for link in soup_results:
                img_urls.append(link.get('src'))
            # http://img1.doubanio.com/view/photo/photo/public/p2194607728.jpg
            i = 1
            for img_url in img_urls:
                image_url = img_url.replace('https', 'http').replace('albumicon', 'photo')
                image_name = title + str(i) + '.jpg'
                i += 1
                self.crawler_image(image_url, image_name)
        else:
            pass

    def crawler_image(self, image_url, image_name):
        time.sleep(np.random.rand()*5)
        try:
            response = requests.get(image_url, headers=self.headers[np.random.randint(0, len(self.headers) - 1)], stream=True)
            with open('images/movies/' + image_name, 'wb') as out_file:
                shutil.copyfileobj(response.raw, out_file)
            del response
        except (requests.exceptions.HTTPError, requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            print(e, image_url)

    def run(self):
        sorted_movie_lists = []
        for topic in self.topics:
            topic_id, topic_name = topic
            movie_list = self.crawler(topic_id, urllib.parse.quote(topic_name))
            movie_list = sorted(movie_list, key=lambda x: (float(x[1]), int(x[2])), reverse=True)
            sorted_movie_lists.append(movie_list)
        tags = [topic[1] for topic in self.topics]
        self.insert_to_excel(sorted_movie_lists, tags)
        print('Finish!!!')

    def insert_to_excel(self, results, tags):
        wb = Workbook(optimized_write=True)
        ws = []
        for i in range(len(tags)):
            ws.append(wb.create_sheet(title=tags[i]))
        for i in range(len(tags)):
            ws[i].append(['序号', '电影名', '评分', '评价人数', '制片地区/类型/上映年份/导演/主演', '链接'])
            count = 1
            for movie_list in results[i]:
                ws[i].append([count, movie_list[0], float(movie_list[1]), int(movie_list[2]), movie_list[3], movie_list[4]])
                count += 1
        save_path = 'movie_lists'
        for i in range(len(tags)):
            save_path += ('-' + tags[i])
        save_path += '.xlsx'
        wb.save(save_path)

    def exist_file(self, search_path, filename):
        candidate = os.path.join(search_path, filename)
        if os.path.isfile(candidate):
            return 'yes'
        return None

if __name__ == '__main__':
    # DoubanMoviesCrawler([['76168', '吉濑美智子']]).run()
    DoubanMoviesCrawler([['73551', '1936'], ['73808', '1935'], ['73956', '1934'], ['73720', '1933'], ['74519', '1932'], ['73917', '1931'], ['70304', '1930'], ['73726', '1929'], ['74226', '1928'], ['74046', '1927'], ['73907', '1926'], ['74349', '1925'], ['74686', '1924'], ['76079', '1923'], ['74685', '1922'], ['74464', '1921'], ['73461', '1920'], ['75146', '1919'], ['78678', '1918'], ['78679', '1917'], ['77850', '1916'], ['75543', '1915'], ['75069', '1914'], ['77405', '1913'], ['77214', '1912'], ['79944', '1911'], ['75113', '1910'], ['74477', '1900']]).run()
    DoubanMoviesCrawler([['72823', '1976'], ['72864', '1975'], ['72830', '1974'], ['72802', '1973'], ['73120', '1972'], ['72762', '1971'], ['72718', '1970'], ['72986', '1969'], ['72822', '1968'], ['72727', '1967'], ['72740', '1966'], ['72925', '1965'], ['72808', '1964'], ['73150', '1963'], ['72825', '1962'], ['72990', '1961'], ['72763', '1960'], ['73207', '1959'], ['73156', '1958'], ['72922', '1957'], ['72894', '1956'], ['73095', '1955'], ['73134', '1954'], ['73197', '1953'], ['72800', '1952'], ['73107', '1951'], ['73019', '1950'], ['73262', '1949'], ['73335', '1948'], ['73228', '1947'], ['73585', '1946'], ['73522', '1945'], ['73989', '1944'], ['74036', '1943'], ['73789', '1942'], ['73739', '1941'], ['73388', '1940'], ['73825', '1939'], ['74377', '1938'], ['73604', '1937']]).run()
    DoubanMoviesCrawler([['73160', '2016'], ['65330', '2015'], ['256', '2014'], ['257', '2013'], ['71847', '2012'], ['71850', '2011'], ['71868', '2010'], ['71939', '2009'], ['72005', '2008'], ['72010', '2007'], ['72014', '2006'], ['72020', '2005'], ['72019', '2004'], ['72025', '2003'], ['72030', '2002'], ['72029', '2001'], ['72037', '2000'], ['72040', '1999'], ['72044', '1998'], ['72038', '1997'], ['72073', '1996'], ['72056', '1995'], ['72035', '1994'], ['72081', '1993'], ['72093', '1992'], ['72138', '1991'], ['72117', '1990'], ['72668', '1989'], ['72673', '1988'], ['72715', '1987'], ['72682', '1986'], ['63701', '1985'], ['72704', '1984'], ['72818', '1983'], ['72739', '1982'], ['72747', '1981'], ['72716', '1980'], ['72948', '1979'], ['72858', '1978'], ['72784', '1977']]).run()
    DoubanMoviesCrawler([['60570', '电视剧'], ['220', '日剧'], ['72047', '日剧SP'], ['60445', '美剧'], ['61457', '英剧'], ['60444', '韩剧'], ['63560', '国产剧'], ['65928', '港剧'], ['65388', 'TVB'], ['66926', '台剧'], ['65900', '泰剧']]).run()
    DoubanMoviesCrawler([['62356', '动画'], ['72053', '动画电影'], ['60407', '动漫'], ['71701', '日本动画'], ['72050', 'OVA'], ['72016', '剧场版'], ['62374', '动画短片'], ['72008', '漫画改编']]).run()
    DoubanMoviesCrawler([['133', '心理'], ['61258', '人生'], ['61554', '女性'], ['61738', '家庭'], ['61469', '生活'], ['60406', '时尚'], ['60465', '美食'], ['71849', '人性'], ['62668', '社会'], ['62667', '政治'], ['60552', '宗教'], ['62371', '战争'], ['67674', '灾难'], ['62376', '传记'], ['60334', '纪录片'], ['62370', '短片'], ['66928', '综艺']]).run()
    DoubanMoviesCrawler([['61870', '冒险'], ['70228', '西部'], ['66201', '公路'], ['273', '旅行'], ['60434', '文艺'], ['72027', '歌舞'], ['60460', '音乐'], ['218', '情感'], ['60521', '励志'], ['60515', '梦想'], ['65937', '感动'], ['62377', '感人'], ['62227', '温暖'], ['66081', '温情'], ['60443', '爱情'], ['62383', '浪漫'], ['62800', '友情'], ['61527', '青春'], ['62258', '校园'], ['62221', '童年'], ['62387', '童话']]).run()
    DoubanMoviesCrawler([['62359', '剧情'], ['60454', '科幻'], ['62386', '史诗'], ['62636', '奇幻'], ['62372', '魔幻'], ['65318', '古装'], ['62375', '情色'], ['62849', 'les'], ['62381', '同志'], ['72022', '伦理'], ['62355', '喜剧'], ['209', '搞笑'], ['62360', '动作'], ['62364', '犯罪'], ['62363', '惊悚'], ['60979', '悬疑'], ['62369', '恐怖'], ['62389', 'cult'], ['62378', '暴力'], ['70686', '血腥'], ['62382', '黑帮'], ['61810', '黑色幽默']]).run()



"""
http://www.douban.com/j/tag/items?start=0&limit=1000000&topic_id=256&topic_name=2014&mod=movie

&topic_id=76168&topic_name=吉濑美智子

&topic_id=62359&topic_name=剧情
&topic_id=60454&topic_name=科幻
&topic_id=62386&topic_name=史诗
&topic_id=62636&topic_name=奇幻
&topic_id=62372&topic_name=魔幻
&topic_id=65318&topic_name=古装
&topic_id=62375&topic_name=情色
&topic_id=62849&topic_name=les
&topic_id=62381&topic_name=同志
&topic_id=72022&topic_name=伦理
&topic_id=62355&topic_name=喜剧
&topic_id=209&topic_name=搞笑
&topic_id=62360&topic_name=动作
&topic_id=62364&topic_name=犯罪
&topic_id=62363&topic_name=惊悚
&topic_id=60979&topic_name=悬疑
&topic_id=62369&topic_name=恐怖
&topic_id=62389&topic_name=cult
&topic_id=62378&topic_name=暴力
&topic_id=70686&topic_name=血腥
&topic_id=62382&topic_name=黑帮
&topic_id=61810&topic_name=黑色幽默
&topic_id=61870&topic_name=冒险
&topic_id=70228&topic_name=西部
&topic_id=66201&topic_name=公路
&topic_id=273&topic_name=旅行
&topic_id=60434&topic_name=文艺
&topic_id=72027&topic_name=歌舞
&topic_id=60460&topic_name=音乐
&topic_id=218&topic_name=情感
&topic_id=60521&topic_name=励志
&topic_id=60515&topic_name=梦想
&topic_id=65937&topic_name=感动
&topic_id=62377&topic_name=感人
&topic_id=62227&topic_name=温暖
&topic_id=66081&topic_name=温情
&topic_id=60443&topic_name=爱情
&topic_id=62383&topic_name=浪漫
&topic_id=62800&topic_name=友情
&topic_id=61527&topic_name=青春
&topic_id=62258&topic_name=校园
&topic_id=62221&topic_name=童年
&topic_id=62387&topic_name=童话
&topic_id=133&topic_name=心理
&topic_id=61258&topic_name=人生
&topic_id=61554&topic_name=女性
&topic_id=61738&topic_name=家庭
&topic_id=61469&topic_name=生活
&topic_id=60406&topic_name=时尚
&topic_id=60465&topic_name=美食
&topic_id=71849&topic_name=人性
&topic_id=62668&topic_name=社会
&topic_id=62667&topic_name=政治
&topic_id=60552&topic_name=宗教
&topic_id=62371&topic_name=战争
&topic_id=67674&topic_name=灾难
&topic_id=62376&topic_name=传记
&topic_id=60334&topic_name=纪录片
&topic_id=62370&topic_name=短片
&topic_id=66928&topic_name=综艺


&topic_id=62356&topic_name=动画
&topic_id=72053&topic_name=动画电影
&topic_id=60407&topic_name=动漫
&topic_id=71701&topic_name=日本动画
&topic_id=72050&topic_name=OVA
&topic_id=72016&topic_name=剧场版
&topic_id=62374&topic_name=动画短片
&topic_id=72008&topic_name=漫画改编


&topic_id=60570&topic_name=电视剧
&topic_id=220&topic_name=日剧
&topic_id=72047&topic_name=日剧SP
&topic_id=60445&topic_name=美剧
&topic_id=61457&topic_name=英剧
&topic_id=60444&topic_name=韩剧
&topic_id=63560&topic_name=国产剧
&topic_id=65928&topic_name=港剧
&topic_id=65388&topic_name=TVB
&topic_id=66926&topic_name=台剧
&topic_id=65900&topic_name=泰剧


&topic_id=73160&topic_name=2016
&topic_id=65330&topic_name=2015
&topic_id=256&topic_name=2014
&topic_id=257&topic_name=2013
&topic_id=71847&topic_name=2012
&topic_id=71850&topic_name=2011
&topic_id=71868&topic_name=2010
&topic_id=71939&topic_name=2009
&topic_id=72005&topic_name=2008
&topic_id=72010&topic_name=2007
&topic_id=72014&topic_name=2006
&topic_id=72020&topic_name=2005
&topic_id=72019&topic_name=2004
&topic_id=72025&topic_name=2003
&topic_id=72030&topic_name=2002
&topic_id=72029&topic_name=2001
&topic_id=72037&topic_name=2000
&topic_id=72040&topic_name=1999
&topic_id=72044&topic_name=1998
&topic_id=72038&topic_name=1997
&topic_id=72073&topic_name=1996
&topic_id=72056&topic_name=1995
&topic_id=72035&topic_name=1994
&topic_id=72081&topic_name=1993
&topic_id=72093&topic_name=1992
&topic_id=72138&topic_name=1991
&topic_id=72117&topic_name=1990
&topic_id=72668&topic_name=1989
&topic_id=72673&topic_name=1988
&topic_id=72715&topic_name=1987
&topic_id=72682&topic_name=1986
&topic_id=63701&topic_name=1985
&topic_id=72704&topic_name=1984
&topic_id=72818&topic_name=1983
&topic_id=72739&topic_name=1982
&topic_id=72747&topic_name=1981
&topic_id=72716&topic_name=1980
&topic_id=72948&topic_name=1979
&topic_id=72858&topic_name=1978
&topic_id=72784&topic_name=1977
&topic_id=72823&topic_name=1976
&topic_id=72864&topic_name=1975
&topic_id=72830&topic_name=1974
&topic_id=72802&topic_name=1973
&topic_id=73120&topic_name=1972
&topic_id=72762&topic_name=1971
&topic_id=72718&topic_name=1970
&topic_id=72986&topic_name=1969
&topic_id=72822&topic_name=1968
&topic_id=72727&topic_name=1967
&topic_id=72740&topic_name=1966
&topic_id=72925&topic_name=1965
&topic_id=72808&topic_name=1964
&topic_id=73150&topic_name=1963
&topic_id=72825&topic_name=1962
&topic_id=72990&topic_name=1961
&topic_id=72763&topic_name=1960
&topic_id=73207&topic_name=1959
&topic_id=73156&topic_name=1958
&topic_id=72922&topic_name=1957
&topic_id=72894&topic_name=1956
&topic_id=73095&topic_name=1955
&topic_id=73134&topic_name=1954
&topic_id=73197&topic_name=1953
&topic_id=72800&topic_name=1952
&topic_id=73107&topic_name=1951
&topic_id=73019&topic_name=1950
&topic_id=73262&topic_name=1949
&topic_id=73335&topic_name=1948
&topic_id=73228&topic_name=1947
&topic_id=73585&topic_name=1946
&topic_id=73522&topic_name=1945
&topic_id=73989&topic_name=1944
&topic_id=74036&topic_name=1943
&topic_id=73789&topic_name=1942
&topic_id=73739&topic_name=1941
&topic_id=73388&topic_name=1940
&topic_id=73825&topic_name=1939
&topic_id=74377&topic_name=1938
&topic_id=73604&topic_name=1937
&topic_id=73551&topic_name=1936
&topic_id=73808&topic_name=1935
&topic_id=73956&topic_name=1934
&topic_id=73720&topic_name=1933
&topic_id=74519&topic_name=1932
&topic_id=73917&topic_name=1931
&topic_id=70304&topic_name=1930
&topic_id=73726&topic_name=1929
&topic_id=74226&topic_name=1928
&topic_id=74046&topic_name=1927
&topic_id=73907&topic_name=1926
&topic_id=74349&topic_name=1925
&topic_id=74686&topic_name=1924
&topic_id=76079&topic_name=1923
&topic_id=74685&topic_name=1922
&topic_id=74464&topic_name=1921
&topic_id=73461&topic_name=1920
&topic_id=75146&topic_name=1919
&topic_id=78678&topic_name=1918
&topic_id=78679&topic_name=1917
&topic_id=77850&topic_name=1916
&topic_id=75543&topic_name=1915
&topic_id=75069&topic_name=1914
&topic_id=77405&topic_name=1913
&topic_id=77214&topic_name=1912
&topic_id=79944&topic_name=1911
&topic_id=75113&topic_name=1910
&topic_id=74477&topic_name=1900
"""