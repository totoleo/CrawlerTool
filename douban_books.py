# -*- coding: utf-8 -*-
__author__ = 'hipponensis'

"""标签: 书名, 评分, 评价人数, 作者/译者/出版信息, 链接"""

import sys
import time
import urllib.request
import urllib.error
import urllib.parse
import gzip
import numpy as np
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


class DoubanBooksCrawler(object):

    def __init__(self, topics):
        self.topics = topics
        self.headers = [
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:40.0) Gecko/20100101 Firefox/40.0', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36', 'Accept-Encoding': 'gzip'}
        ]

    def crawler(self, topic_id, topic_name):
        url = 'http://www.douban.com/j/tag/items?start=0&limit=500&topic_id=' + topic_id + '&topic_name=' + topic_name + '&mod=book'
        book_list = []
        try:
            randhds = self.headers[np.random.randint(0, len(self.headers)-1)]
            req = urllib.request.Request(url, headers=randhds)
            resp = urllib.request.urlopen(req)
            try:
                g = gzip.GzipFile(mode="rb", fileobj=resp)
                source_code = g.read()
            except:
                source_code = resp.read()
            source_code = eval(source_code.decode('utf-8', 'ignore'))
            plain_text = str(source_code['html'])
            total = int(source_code['total'])
            plain_text = plain_text.replace('\\', '')
            plain_text = plain_text.replace('\n', '')
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(e, url)
            plain_text = ''
            total = 0

        list_info = re.split(r'<dl>', plain_text)[1:]

        if list_info is None:
            print('no result', url)

        for book_info in list_info:
            soup = BeautifulSoup(book_info, 'html.parser')
            book_url = soup.find('a').get('href').replace('https', 'http')
            title = soup.find('a', class_="title").get_text()
            desc = soup.find('div', class_="desc").get_text()
            desc = desc.strip(' ')
            desc = desc.strip('\n')
            try:
                rating = soup.find('span', class_="rating_nums").get_text()
            except:
                rating = '0.0'
            try:
                if rating != '0.0':
                    votes = self.crawler_votes(book_url)
                    votes = votes.replace('\\xe4\\xba\\xba\\xe8\\xaf\\x84\\xe4\\xbb\\xb7', '')
                else:
                    votes = '0'
            except:
                votes = '0'
            book_list.append([title, rating, votes, desc, book_url])

        if total >= 500:
            for i in range(1, int(total/500) + 1):
                time.sleep(np.random.rand()*5)
                url = 'http://www.douban.com/j/tag/items?start=' + str(i * 500) + '&limit=500&topic_id=' + topic_id + '&topic_name=' + topic_name + '&mod=book'
                try:
                    randhds = self.headers[np.random.randint(0, len(self.headers)-1)]
                    req = urllib.request.Request(url, headers=randhds)
                    resp = urllib.request.urlopen(req)
                    try:
                        g = gzip.GzipFile(mode="rb", fileobj=resp)
                        source_code = g.read()
                    except:
                        source_code = resp.read()
                    if '{' in source_code.decode('utf-8', 'ignore'):
                        source_code = eval(source_code.decode('utf-8', 'ignore'))
                    else:
                        source_code = eval('{"' + source_code.decode('utf-8', 'ignore'))
                    plain_text = str(source_code['html'])
                    plain_text = plain_text.replace('\\', '')
                    plain_text = plain_text.replace('\n', '')
                except (urllib.error.HTTPError, urllib.error.URLError) as e:
                    print(e, url)
                    plain_text = ''

                list_info = re.split(r'<dl>', plain_text)[1:]

                if list_info is None:
                    print('no result', url)

                for book_info in list_info:
                    soup = BeautifulSoup(book_info, 'html.parser')
                    book_url = soup.find('a').get('href').replace('https', 'http')
                    title = soup.find('a', class_="title").get_text()
                    desc = soup.find('div', class_="desc").get_text()
                    desc = desc.strip(' ')
                    desc = desc.strip('\n')
                    try:
                        rating = soup.find('span', class_="rating_nums").get_text()
                    except:
                        rating = '0.0'
                    try:
                        if rating != '0.0':
                            votes = self.crawler_votes(book_url)
                            votes = votes.replace('\\xe4\\xba\\xba\\xe8\\xaf\\x84\\xe4\\xbb\\xb7', '')
                        else:
                            votes = '0'
                    except:
                        votes = '0'
                    book_list.append([title, rating, votes, desc, book_url])
        return book_list

    def crawler_votes(self, book_url):
        time.sleep(np.random.rand()*5)
        try:
            req = urllib.request.Request(book_url, headers=self.headers[np.random.randint(0, len(self.headers) - 1)])
            resp = urllib.request.urlopen(req)
            try:
                g = gzip.GzipFile(mode="rb", fileobj=resp)
                source_code = g.read()
            except:
                source_code = resp.read()
            plain_text = str(source_code)
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(e, book_url)
            plain_text = ''
        soup = BeautifulSoup(plain_text, 'html.parser')
        votes = soup.find(id="interest_sectl").find('a', class_="rating_people").get_text()
        return votes

    def run(self):
        sorted_book_lists = []
        for topic in self.topics:
            topic_id, topic_name = topic
            book_list = self.crawler(topic_id, urllib.parse.quote(topic_name))
            book_list = sorted(book_list, key=lambda x: (float(x[1]), int(x[2])), reverse=True)
            sorted_book_lists.append(book_list)
        tags = [topic[1] for topic in self.topics]
        self.insert_to_excel(sorted_book_lists, tags)
        print('Finish!!!')

    def insert_to_excel(self, results, tags):
        wb = Workbook(optimized_write=True)
        ws = []
        for i in range(len(tags)):
            ws.append(wb.create_sheet(title=tags[i]))
        for i in range(len(tags)):
            ws[i].append(['序号', '书名', '评分', '评价人数', '作者/译者/出版社/出版时间/价格', '链接'])
            count = 1
            for booklist in results[i]:
                try:
                    book_list = []
                    for each in booklist:
                        each = ILLEGAL_CHARACTERS_RE.sub(r'', each)
                        book_list.append(each)
                    if book_list:
                        ws[i].append([count, book_list[0], float(book_list[1]), int(book_list[2]), book_list[3], book_list[4]])
                        count += 1
                    else:
                        print('Fail to save i:', i, book_list)
                except:
                    print('Fail to save i:', i, booklist)
                    pass
        save_path = 'book_lists'
        for i in range(len(tags)):
            save_path += ('-' + tags[i])
        save_path += '.xlsx'
        wb.save(save_path)

if __name__ == '__main__':
    DoubanBooksCrawler([['71843', '工具书'], ['71840', '教材']]).run()
    # DoubanBooksCrawler([['60494', '科学'], ['62726', '科普'], ['60512', '科技'], ['62672', '数学'], ['65403', '计算机'], ['62732', '算法'], ['62155', '编程'], ['60461', '程序'], ['62733', 'web'], ['62737', '通信'], ['60502', '互联网'], ['62739', '神经网络'], ['62738', '交互'], ['62730', '交互设计'], ['62731', '用户体验'], ['62735', 'UE'], ['62736', 'UCD']]).run()
    # DoubanBooksCrawler([['61431', '经济'], ['62713', '经济学'], ['62724', '企业史'], ['62717', '商业'], ['61640', '管理'], ['62719', '营销'], ['60497', '创业'], ['61432', '金融'], ['60493', '投资'], ['139', '理财'], ['210', '股票'], ['62722', '广告'], ['62725', '策划']]).run()
    # DoubanBooksCrawler([['61258', '人生'], ['61469', '生活'], ['67703', '思维'], ['218', '情感'], ['133', '心理'], ['60418', '心理学'], ['62212', '个人管理'], ['62711', '人际关系'], ['260', '职场'], ['60521', '励志'], ['60487', '成长'], ['61474', '教育'], ['61554', '女性'], ['60919', '两性'], ['216', '健康'], ['62704', '灵修'], ['61204', '养生'], ['60465', '美食'], ['273', '旅行'], ['62712', '自助游'], ['60534', '家居'], ['95', '手工']]).run()
    # DoubanBooksCrawler([['60403', '艺术'], ['93', '艺术史'], ['219', '建筑'], ['60460', '音乐'], ['61256', '戏剧'], ['60476', '绘画'], ['60475', '美术'], ['105', '书法'], ['60473', '设计'], ['272', '摄影'], ['60459', '电影'], ['62685', '军事']]).run()
    # DoubanBooksCrawler([['60501', '文化'], ['60489', '历史'], ['62677', '中国历史'], ['62689', '近代史'], ['62690', '考古'], ['60663', '人文'], ['62376', '传记'], ['62674', '回忆录'], ['62667', '政治'], ['62673', '政治学'], ['60552', '宗教'], ['61608', '哲学'], ['62675', '思想'], ['62072', '国学'], ['62668', '社会'], ['62664', '社会学'], ['63999', '人类学']]).run()
    # DoubanBooksCrawler([['255', '小说'], ['60454', '科幻'], ['62636', '奇幻'], ['62630', '武侠'], ['62625', '推理'], ['60979', '悬疑'], ['60443', '爱情'], ['61527', '青春'], ['62627', '言情'], ['62640', '网络小说'], ['62644', '轻小说'], ['62641', '穿越'], ['60918', '耽美'], ['60429', '漫画'], ['98', '绘本']]).run()
    # DoubanBooksCrawler([['60456', '文学'], ['62597', '外国文学'], ['62600', '中国文学'], ['62612', '古典文学'], ['62358', '经典'], ['62613', '名著'], ['254', '随笔'], ['60509', '散文'], ['61504', '诗歌'], ['62618', '诗词'], ['61747', '杂文'], ['60519', '游记']]).run()


"""

http://www.douban.com/j/tag/items?start=0&limit=17592&topic_id=254&topic_name=%E9%9A%8F%E7%AC%94&mod=book

limit=200959&topic_id=60456&topic_name=文学
limit=32131&topic_id=62597&topic_name=外国文学
limit=23037&topic_id=62600&topic_name=中国文学
limit=7117&topic_id=62612&topic_name=古典文学
limit=15182&topic_id=62358&topic_name=经典
limit=3078&topic_id=62613&topic_name=名著
limit=17592&topic_id=254&topic_name=随笔
limit=14292&topic_id=60509&topic_name=散文
limit=8534&topic_id=61504&topic_name=诗歌
limit=3101&topic_id=62618&topic_name=诗词
limit=3338&topic_id=61747&topic_name=杂文
limit=3170&topic_id=60519&topic_name=游记

limit=77025&topic_id=255&topic_name=小说
limit=8049&topic_id=60454&topic_name=科幻
limit=6809&topic_id=62636&topic_name=奇幻
limit=4026&topic_id=62630&topic_name=武侠
limit=10271&topic_id=62625&topic_name=推理
limit=6819&topic_id=60979&topic_name=悬疑
limit=12680&topic_id=60443&topic_name=爱情
limit=8609&topic_id=61527&topic_name=青春
limit=9516&topic_id=62627&topic_name=言情
limit=5370&topic_id=62640&topic_name=网络小说
limit=6653&topic_id=62644&topic_name=轻小说
limit=2200&topic_id=62641&topic_name=穿越
limit=9289&topic_id=60918&topic_name=耽美
# 网文不爬了...男频yousuu网...女频未知
limit=47933&topic_id=60429&topic_name=漫画
limit=17068&topic_id=98&topic_name=绘本


limit=25638&topic_id=60501&topic_name=文化
limit=65279&topic_id=60489&topic_name=历史
limit=4705&topic_id=62677&topic_name=中国历史
limit=3181&topic_id=62689&topic_name=近代史
limit=3342&topic_id=62690&topic_name=考古
limit=91291&topic_id=60663&topic_name=人文
limit=19526&topic_id=62376&topic_name=传记
limit=3485&topic_id=62674&topic_name=回忆录
limit=10189&topic_id=62667&topic_name=政治
limit=8837&topic_id=62673&topic_name=政治学
limit=16195&topic_id=60552&topic_name=宗教
limit=22619&topic_id=61608&topic_name=哲学
limit=7862&topic_id=62675&topic_name=思想
limit=6085&topic_id=62072&topic_name=国学
limit=23961&topic_id=62668&topic_name=社会
limit=14187&topic_id=62664&topic_name=社会学
limit=4476&topic_id=63999&topic_name=人类学

limit=29796&topic_id=60403&topic_name=艺术
limit=4405&topic_id=93&topic_name=艺术史
limit=14160&topic_id=219&topic_name=建筑
limit=4834&topic_id=60460&topic_name=音乐
limit=3316&topic_id=61256&topic_name=戏剧
limit=8833&topic_id=60476&topic_name=绘画
limit=2310&topic_id=60475&topic_name=美术
limit=2224&topic_id=105&topic_name=书法
limit=16894&topic_id=60473&topic_name=设计
limit=9413&topic_id=272&topic_name=摄影
limit=5366&topic_id=60459&topic_name=电影
limit=3652&topic_id=62685&topic_name=军事


limit=18046&topic_id=61258&topic_name=人生
limit=14297&topic_id=61469&topic_name=生活
limit=5526&topic_id=67703&topic_name=思维
limit=22788&topic_id=218&topic_name=情感
limit=25613&topic_id=133&topic_name=心理
limit=17059&topic_id=60418&topic_name=心理学
limit=2525&topic_id=62212&topic_name=个人管理
limit=1219&topic_id=62711&topic_name=人际关系
limit=5512&topic_id=260&topic_name=职场
limit=7391&topic_id=60521&topic_name=励志
limit=16896&topic_id=60487&topic_name=成长
limit=14697&topic_id=61474&topic_name=教育
limit=8125&topic_id=61554&topic_name=女性
limit=975&topic_id=60919&topic_name=两性
limit=4836&topic_id=216&topic_name=健康
limit=2902&topic_id=62704&topic_name=灵修
limit=1682&topic_id=61204&topic_name=养生
limit=4474&topic_id=60465&topic_name=美食
limit=9461&topic_id=273&topic_name=旅行
limit=144&topic_id=62712&topic_name=自助游
limit=595&topic_id=60534&topic_name=家居
limit=3011&topic_id=95&topic_name=手工



&topic_id=61431&topic_name=经济
&topic_id=62713&topic_name=经济学
&topic_id=62724&topic_name=企业史
&topic_id=62717&topic_name=商业
&topic_id=61640&topic_name=管理
&topic_id=62719&topic_name=营销
&topic_id=60497&topic_name=创业
&topic_id=61432&topic_name=金融
&topic_id=60493&topic_name=投资
&topic_id=139&topic_name=理财
&topic_id=210&topic_name=股票
&topic_id=62722&topic_name=广告
&topic_id=62725&topic_name=策划

&topic_id=60494&topic_name=科学
&topic_id=62726&topic_name=科普
&topic_id=60512&topic_name=科技
&topic_id=62672&topic_name=数学
&topic_id=62732&topic_name=算法
&topic_id=62155&topic_name=编程
&topic_id=60461&topic_name=程序
&topic_id=62733&topic_name=web
&topic_id=62737&topic_name=通信
&topic_id=60502&topic_name=互联网
&topic_id=62739&topic_name=神经网络
&topic_id=62738&topic_name=交互
&topic_id=62730&topic_name=交互设计
&topic_id=62731&topic_name=用户体验
&topic_id=62735&topic_name=UE
&topic_id=62736&topic_name=UCD

limit=7841&topic_id=71843&topic_name=工具书
limit=10634&topic_id=71840&topic_name=教材

"""





"""为什么代理IP爬太快了也全部403???


import sys
import time
import urllib.request
import urllib.error
import urllib.parse
import gzip
import numpy as np
import redis
import re
from bs4 import BeautifulSoup
from multiprocessing.dummy import Pool as ThreadPool
from openpyxl import Workbook


class DoubanBooksCrawler(object):

    def __init__(self, topics):
        self.topics = topics
        self.headers = [
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:40.0) Gecko/20100101 Firefox/40.0', 'Accept-Encoding': 'gzip'},
            {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36', 'Accept-Encoding': 'gzip'}
        ]
        self.pool = ThreadPool(20)
        self.count = 0
        self.conn = redis.Redis(host='127.0.0.1', port=6379, db=0)
        self.ip = self.conn.zrange('ipproxy:3', 0, -1)
        self.ip = self.ip[0:(len(self.ip) - 20)]

    def crawler(self, topic_id, topic_name):
        referer = 'http://www.douban.com/tag/' + topic_name + '/?focus=book'
        hds = [
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:34.0) Gecko/20100101 Firefox/34.0',
             'Accept-Encoding': 'gzip, deflate, sdch', 'referer': referer, 'x-requested-with': 'XMLHttpRequest'},
            {'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6',
             'Accept-Encoding': 'gzip, deflate, sdch', 'referer': referer, 'x-requested-with': 'XMLHttpRequest'},
            {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11',
             'Accept-Encoding': 'gzip, deflate, sdch', 'referer': referer, 'x-requested-with': 'XMLHttpRequest'},
            {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)',
             'Accept-Encoding': 'gzip, deflate, sdch', 'referer': referer, 'x-requested-with': 'XMLHttpRequest'},
            {'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:40.0) Gecko/20100101 Firefox/40.0',
             'Accept-Encoding': 'gzip, deflate, sdch', 'referer': referer, 'x-requested-with': 'XMLHttpRequest'},
            {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/44.0.2403.89 Chrome/44.0.2403.89 Safari/537.36',
             'Accept-Encoding': 'gzip, deflate, sdch', 'referer': referer, 'x-requested-with': 'XMLHttpRequest'}
        ]
        url = 'http://www.douban.com/j/tag/items?start=0&limit=1000000&topic_id=' + topic_id + '&topic_name=' + topic_name + '&mod=book'
        randipnum = np.random.randint(0, len(self.ip)-1)
        proxy_ip = self.ip[randipnum].decode('utf-8')
        proxy_support = urllib.request.ProxyHandler({'http': proxy_ip})
        opener = urllib.request.build_opener(proxy_support)
        urllib.request.install_opener(opener)
        try:
            randhds = hds[np.random.randint(0, len(hds)-1)]
            req = urllib.request.Request(url, headers=randhds)
            resp = urllib.request.urlopen(req)
            try:
                g = gzip.GzipFile(mode="rb", fileobj=resp)
                source_code = g.read()
            except:
                source_code = resp.read()
            source_code = eval(source_code.decode('utf-8', 'ignore'))
            plain_text = str(source_code['html'])
            total = int(source_code['total'])
            plain_text = plain_text.replace('\\', '')
            plain_text = plain_text.replace('\n', '')
            list_info = re.split(r'<dl>', plain_text)[1:]
            if list_info is None or len(list_info) <= 1:
                print(list_info)
            book_list = self.pool.map(self.crawler_info, list_info)
            return book_list
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(e, topic_id, topic_name, proxy_ip)
            plain_text = ''
            self.conn.zrem('ipproxy:1', proxy_ip)
            del self.ip[randipnum]
            self.crawler(topic_id, topic_name)

    def crawler_info(self, book_info):
        soup = BeautifulSoup(book_info, 'html.parser')
        book_url = soup.find('a').get('href')
        title = soup.find('a', class_="title").get_text()
        desc = soup.find('div', class_="desc").get_text()
        desc = desc.strip(' ')
        desc = desc.strip('\n')
        try:
            rating = soup.find('span', class_="rating_nums").get_text()
        except:
            rating = '0.0'
        try:
            if rating != '0.0':
                votes = self.crawler_votes(book_url)
                votes = votes.replace('\\xe4\\xba\\xba\\xe8\\xaf\\x84\\xe4\\xbb\\xb7', '')
            else:
                votes = '0'
        except:
            votes = '0'
        return [title, rating, votes, desc, book_url]

    def crawler_votes(self, book_url):
        if self.count >= len(self.ip):
            self.count = 0
        proxy_ip = (self.ip[self.count]).decode('utf-8')
        self.count += 1
        try:
            proxy_support = urllib.request.ProxyHandler({'http': proxy_ip})
            opener = urllib.request.build_opener(proxy_support)
            urllib.request.install_opener(opener)
            req = urllib.request.Request(book_url, headers=self.headers[np.random.randint(0, len(self.headers) - 1)])
            resp = urllib.request.urlopen(req)
            try:
                g = gzip.GzipFile(mode="rb", fileobj=resp)
                source_code = g.read()
            except:
                source_code = resp.read()
            plain_text = str(source_code)
            soup = BeautifulSoup(plain_text, 'html.parser')
            votes = soup.find(id="interest_sectl").find('a', class_="rating_people").get_text()
            return votes
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            if e.code != 404:
                print(e, book_url, 'fail ip: %s' % proxy_ip)
                self.conn.zrem('ipproxy:1', proxy_ip)
                del self.ip[self.count - 1]
                self.crawler_votes(book_url)
            else:
                print(e, book_url)

    def run(self):
        sorted_book_lists = []
        for topic in self.topics:
            topic_id, topic_name = topic
            book_list = self.crawler(topic_id, urllib.parse.quote(topic_name))
            book_list = sorted(book_list, key=lambda x: (float(x[1]), int(x[2])), reverse=True)
            sorted_book_lists.append(book_list)
        tags = [topic[1] for topic in self.topics]
        self.insert_to_excel(sorted_book_lists, tags)
        print('Finish!!!')

    def insert_to_excel(self, results, tags):
        wb = Workbook(optimized_write=True)
        ws = []
        for i in range(len(tags)):
            ws.append(wb.create_sheet(title=tags[i]))
        for i in range(len(tags)):
            ws[i].append(['序号', '书名', '评分', '评价人数', '作者/译者/出版社/出版时间/价格', '链接'])
            count = 1
            for book_list in results[i]:
                ws[i].append([count, book_list[0], float(book_list[1]), int(book_list[2]), book_list[3], book_list[4]])
                count += 1
        save_path = 'book_lists'
        for i in range(len(tags)):
            save_path += ('-' + tags[i])
        save_path += '.xlsx'
        wb.save(save_path)

if __name__ == '__main__':
    # DoubanBooksCrawler([['71843', '工具书'], ['71840', '教材']]).run()
    # DoubanBooksCrawler([['60494', '科学'], ['62726', '科普'], ['60512', '科技'], ['62672', '数学'], ['65403', '计算机'], ['62732', '算法'], ['62155', '编程'], ['60461', '程序'], ['62733', 'web'], ['62737', '通信'], ['60502', '互联网'], ['62739', '神经网络'], ['62738', '交互'], ['62730', '交互设计'], ['62731', '用户体验'], ['62735', 'UE'], ['62736', 'UCD']]).run()
    DoubanBooksCrawler([['61431', '经济'], ['62713', '经济学'], ['62724', '企业史'], ['62717', '商业'], ['61640', '管理'], ['62719', '营销'], ['60497', '创业'], ['61432', '金融'], ['60493', '投资'], ['139', '理财'], ['210', '股票'], ['62722', '广告'], ['62725', '策划']]).run()
    DoubanBooksCrawler([['61258', '人生'], ['61469', '生活'], ['67703', '思维'], ['218', '情感'], ['133', '心理'], ['60418', '心理学'], ['62212', '个人管理'], ['62711', '人际关系'], ['260', '职场'], ['60521', '励志'], ['60487', '成长'], ['61474', '教育'], ['61554', '女性'], ['60919', '两性'], ['216', '健康'], ['62704', '灵修'], ['61204', '养生'], ['60465', '美食'], ['273', '旅行'], ['62712', '自助游'], ['60534', '家居'], ['95', '手工']]).run()
    DoubanBooksCrawler([['60403', '艺术'], ['93', '艺术史'], ['219', '建筑'], ['60460', '音乐'], ['61256', '戏剧'], ['60476', '绘画'], ['60475', '美术'], ['105', '书法'], ['60473', '设计'], ['272', '摄影'], ['60459', '电影'], ['62685', '军事']]).run()
    DoubanBooksCrawler([['60501', '文化'], ['60489', '历史'], ['62677', '中国历史'], ['62689', '近代史'], ['62690', '考古'], ['60663', '人文'], ['62376', '传记'], ['62674', '回忆录'], ['62667', '政治'], ['62673', '政治学'], ['60552', '宗教'], ['61608', '哲学'], ['62675', '思想'], ['62072', '国学'], ['62668', '社会'], ['62664', '社会学'], ['63999', '人类学']]).run()
    DoubanBooksCrawler([['255', '小说'], ['60454', '科幻'], ['62636', '奇幻'], ['62630', '武侠'], ['62625', '推理'], ['60979', '悬疑'], ['60443', '爱情'], ['61527', '青春'], ['62627', '言情'], ['62640', '网络小说'], ['62644', '轻小说'], ['62641', '穿越'], ['60918', '耽美'], ['60429', '漫画'], ['98', '绘本']]).run()
    DoubanBooksCrawler([['60456', '文学'], ['62597', '外国文学'], ['62600', '中国文学'], ['62612', '古典文学'], ['62358', '经典'], ['62613', '名著'], ['254', '随笔'], ['60509', '散文'], ['61504', '诗歌'], ['62618', '诗词'], ['61747', '杂文'], ['60519', '游记']]).run()

"""